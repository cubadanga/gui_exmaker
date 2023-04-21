from colorama import init, Fore
import numpy as np
import pandas as pd
import openpyxl
from  openpyxl.styles  import  Font
import random
import re
import os
import sys
import shutil
import time
from urllib.error import HTTPError
import urllib.request
from urllib.request import urlopen
from urllib.parse import urlparse, parse_qs
import configparser
from bs4 import BeautifulSoup
import webbrowser

#프린트문 색상 변경을 위해 초기화
np.set_printoptions(threshold=np.inf, linewidth=np.inf)
init()
print(Fore.LIGHTBLUE_EX + "엑셀파일 작성을 시작 합니다. 작성중..." )
print(Fore.RESET)

def loadPassword(): #우선 'set.ini' 파일에 저장된 패스워드와 웹에 있는 패스워드가 일치하는지 확인한다.
    basedir = os.getcwd()
    ini_dir = os.path.join(basedir,'set.ini')

    # pc set.ini 파일의 저장된 pass워드 읽어오기
    properties = configparser.ConfigParser()
    properties.read(ini_dir)
    
    if 'DEFAULT' in properties and 'userpass' in properties['DEFAULT']:
        password = properties['DEFAULT']['userpass']
        return password
    else:
        print(Fore.RED + "오류 - 'userpass' key not found in set.ini file."+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def getPtag(url): # 웹페이지에 적어 놓은 password 텍스트를 크롤링해 추출하는 함수
    try:
        html = urlopen(url)
        
    except HTTPError as e:
        print(Fore.RED + '오류 - 네트워크오류 또는 패스워드url오류'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
    try:
        soup = BeautifulSoup(html,"html.parser")
        ptag = soup.find('p')
        
    except AttributeError as e:
        return None
    return ptag.text

def judge(password,passTag): #set.ini에 저장된 패스워드와 웹에 있는 패스워드를 비교하는 함수.
    if password == passTag:
        properties = configparser.ConfigParser()
        properties.set('DEFAULT','userpass',password)
        with open('./set.ini','w',encoding='utf-8') as F:
            properties.write(F)
                
        print("이번 달 패스워드 확인 완료! 오늘도 파이팅!")
        pass
    else:
        print(Fore.RED + "오류 - 저장된 패스워드가 없거나 올바른 패스워드가 아닙니다. 패스워드는 단체방 금월 암호 공지를 확인하세요."+Fore.RESET+'\n')
        inputPass(password,passTag)

def inputPass(password,passTag): #패스워드가 틀렸을 때 콘솔에서 다시 입력을 받는 함수
    userPass = str(password)
    passTag = passTag
    print('\n' + "패스워드를 입력해 주세요.")
    userPass = input()
    judge(userPass, passTag)

def readExcel(product_path, setting_path): #셋팅 데이터와 유저가 입력한 제품 데이터를 읽어 온다.
    try:
        writeSheet_DF = pd.read_excel(product_path, sheet_name = 'write', header = 0)
        setting_DF = pd.read_excel(setting_path, sheet_name = 'setting', header = 0)
        setting_DF = setting_DF.fillna('')

        return writeSheet_DF, setting_DF
        
    except ValueError as e:
        print(Fore.RED + '오류 - 엑셀 시트의 시트명이 다르거나 올바른 파일이 아닙니다.'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    except FileNotFoundError as e:
        print(Fore.RED + '오류 - product.xlsx 파일을 찾을 수 없습니다.'+'\n'+'이런 경우, 파일명이 잘못된 경우가 대부분이었습니다.'+' 이 파일은 필수 파일입니다.'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def extract_id(site, url): #쇼핑몰에 따라 제품 ID를 크롤링하고 쇼핑몰에 따라 제품 url을 생성한다.
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    
    if site == 'taobao':
        product_id = query_params.get('id', [''])[0]
        product_url = "https://item.taobao.com/item.htm?id=" + product_id
        return product_id, product_url
    
    elif site == 'shop1688':
        file_name = os.path.splitext(os.path.basename(parsed_url.path))[0]
        product_id = file_name.split("_")[-1]
        product_url = "https://detail.1688.com/offer/" + product_id + ".html"
        return product_id, product_url
    
    elif site == 'vvic':
        file_name = os.path.splitext(os.path.basename(parsed_url.path))[0]
        product_id = file_name.split("_")[-1]
        product_url = "https://www.vvic.com/item/" + product_id
        return product_id, product_url
    
    elif site == 'aliexpress':
        file_name = os.path.splitext(os.path.basename(parsed_url.path))[0]
        product_id = file_name.split("_")[-1]
        product_url = "https://ko.aliexpress.com/item/" + product_id + ".html"
        return product_id, product_url
    
    else:
        return "", ""

def progress_text(productCord, videourl): #데이터 추출관련 문구 출력
    if productCord =="":
        print(Fore.RED + '오류 - 입력한 주소가 해당 쇼핑몰의 주소인지 확인하세요. \n예) 타오바오는 "taobao", 1688은 "shop1688"이라고 입력하셔야 합니다.'+Fore.RESET+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    else:
        print(f'1. 입력 사이트: [{shop_type}]')
        print(f'2. 사이트 url 추출성공: [{product_url}]')
        print(f'3. 제품코드 추출성공: [{productCord}]')
        print(f'4. 제목 추출 성공: [{pName}]')

    if videourl == 'nan':
        videourl = '동영상이 없습니다.'
        print('5. 동영상 url은 없었습니다.')
        
    else:
        print('5. 동영상 url 복사완료!')

def optionTitle(write_df): #네이버 포멧 옵션명 제작 함수
    df = write_df
    try:
        df_goods = df.iloc[0:,5:7]
        df_goods.replace('', np.nan, inplace=True)
        goods_Tclear = df_goods.dropna(axis=1)
        return str("\n".join(goods_Tclear.columns))
        
    except KeyError:
        print(Fore.RED + '오류 - 옵션 금액 또는 옵션내용이 잘못 기입 되었습니다.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def opImg_Download(op_imgurls): #옵션 이미지를 폴더에 다운로드 함.
    try:
        optionNum = 0
        for i in op_imgurls: 
            file_ext = i.split('.')[-1] # 확장자 추출
            path = path_Option + '/' + productCord + '_option_' + str(optionNum)+'.' + file_ext
            random_number = round(random.uniform(0.07, 0.2), 2)
            
            time.sleep(random_number)
            urllib.request.urlretrieve(i, path)
            print(Fore.GREEN + str(optionNum)+'번 옵션 이미지 다운로드 성공'+Fore.RESET)
            optionNum +=1
            
    except urllib.error.HTTPError:
        print(Fore.RED + '오류 - 크롬 브라우저로 타오바오에 로그인이 필요하거나 올바른 옴션 url이 아닙니다.')
        print(Fore.RESET + str(optionNum)+'번 상세 이미지주소: ',i)
        print("엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def createFolder(directory): # 결과 파일 저장 폴더 생성
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print (Fore.RED + '오류 - Creating directory. ' +  directory)
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def descImg_Download(descPages): # 상세 이미지 url 추출 및 다운로드
    descPages = descPages.replace('?getAvatar=avatar','')
    modUrls = re.findall('<img.*?src=[\'"](.*?)[\'"].*?>', descPages)

    try:    
        for i in modUrls:
            descimgNum = 0
            file_ext = i.split('.')[-1] # 확장자 추출
            path = path_Desc + '/' + productCord + '_desc_' + str(descimgNum)+'.' + file_ext
            random_number = round(random.uniform(0.02, 0.3), 2)
            time.sleep(random_number)
            urllib.request.urlretrieve(i, path)
            print(Fore.GREEN +  str(descimgNum)+'번 상세 이미지 다운로드 성공' + Fore.RESET)
            descimgNum +=1

    except urllib.error.HTTPError:
        print(Fore.RED + '오류 - 해외쇼핑몰 로그인이 필요하거나 올바른 상세 url이 아닙니다.')
        print(str(descimgNum)+'번 오류 상세 이미지주소: ',i)
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    except urllib.error.URLError:
        print(Fore.RED + '오류 - 올바른 상세 url이 아닙니다.')
        print('오류 있는 '+str(descimgNum)+'번째 상세 이미지 주소: ',i,'\n(url을 콘트롤키+클릭하면 브라우저에서 오픈합니다.)\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def write_categori_name(categori_num): #카테고리 번호를 넣으면 네이버 카테고리 이름을 찾아 경로를 한글로 써준다.
    naver_Categori = pd.read_excel('./product.xlsx', sheet_name = 'categori_naver', header = 0) #네이버 카테고리 데이터가 저장된 위치
    df_cat = naver_Categori.loc[naver_Categori['카테고리번호'] == categori_num].fillna("")
    strCalevel1 = df_cat['대분류'].to_string(index=False)
    strCalevel2 = df_cat['중분류'].to_string(index=False)
    strCalevel3 = df_cat['소분류'].to_string(index=False)
    strCalevel4 = df_cat['세분류'].to_string(index=False)
    return strCalevel1, strCalevel2, strCalevel3, strCalevel4


def mainImg_Edit(file_path, output_path): #특정 폴더에 담겨있는 메인이미지를 불러와서 product 코드로 이름을 변경하고 제일 첫번째를 메인이미지로 나머지를 서브이미지로 엑셀파일을 작성함.
    try:
        file_names = os.listdir(file_path) #file_path가 유저가 선택한 폴더의 경로

    except FileNotFoundError as e:
        print(Fore.RED + '오류 - mainImage(메인이미지) 폴더가 존재하지 않습니다.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    if len(file_names) > 0:
        i = 1
        j = 0
        images=[]

        for name in file_names:
            src = os.path.join(file_path,name)
            dst = productCord + '-' + str(i) + '.jpg'
            images.append(dst)
            dst = os.path.join(output_path,dst)
            os.rename(src,dst)
            i += 1
            j += 1

        mainImage = images[0]
        del images[0]
        subImages = ",".join(images)
        #destination = "./excel/images/" + productCord + '-' + str(i) + '.jpg'
           
        file_dir = os.path.dirname(file_path+'/')
        file_cnt = 1
        for path, dirs, files in os.walk(file_dir):
            for file in files:
                rename_file_path = os.path.join(path,file)
                file_cnt += 1
                dest_path = './excel/' + file
                shutil.copy(rename_file_path, dest_path)
                
    else:
        print(Fore.RED + "오류 - 메인이미지 폴더에 이미지가 없습니다.")
        mainImage = ""
        subImages = ""
    print(Fore.RESET + "9. 메인이미지 이름변경/폴더이동 완료!")
    return mainImage, subImages

def price_Calculation(writeSheet_DF):
    df_optiongoods = writeSheet_DF.iloc[0:,5:11]
    df_optiongoods.replace('', np.nan, inplace=True)
    goods_clear = df_optiongoods.dropna(axis=1).copy()
    option_gooddf = goods_clear.columns
    optionColcnt = len(goods_clear.columns)

    optionT1 = [] #옵션1의 옵션제목과 내용
    optionT2 = [] #옵션2의 옵션제목과 내용

    if optionColcnt == 6:
        optionT1 = option_gooddf[0]
        optionT2 = option_gooddf[1]

    elif optionColcnt == 5:
        optionT1 = option_gooddf[0]

    try:
        # 결제 통화 셋팅
        currency_type = goods_clear['결제통화'][0]

        if currency_type =='CNY':
            rate = rate_CNY
            payment_fee = 1.03
            duty = round(goods_clear['물건가격'].max() * payment_fee * int(rate)/rate_USD)
        elif currency_type =='USD':
            rate = rate_USD
            duty = round(goods_clear['물건가격'].max())
        else:
            rate = 1
            duty = round(goods_clear['물건가격'].max()/rate_USD)

    ### 기본 판매가 계산(옵션별 판매가격 계산)
    # * 구매원가 = (상품가(상품가*수수료*환율)+배송비) prime_cost
    # * 기본판매가 = 구매원가*가중치 price_min
    # * 마진 = 기본가-스토어수수료-상품가-배송비 
    # * 마진율 = 마진금액/기본가

        goods_clear['구매원가'] = goods_clear['물건가격'] * payment_fee * int(rate) + goods_clear['실제배송비']
        goods_clear['기본판매가'] = goods_clear['구매원가']*fomul
        prime_cost = round(goods_clear['구매원가'].min())
        
    except KeyError:
        print(Fore.RED + '옵션입력오류 - 옵션1, 옵션2에 옵션명은 적었으나 옵션항목을 입력하지 않았거나\n또는 입력한 옵션의 금액, 화폐단위, 재고수량, 배송비 중 중간에 입력하지 않은 셀이 있는지 보세요.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    # ==============================================#####

    goods_clear['마진'] = goods_clear['기본판매가']-(goods_clear['기본판매가']*fee_naver/100)-goods_clear['구매원가']
    goods_clear['마진율'] = round(goods_clear['마진']/goods_clear['기본판매가']*100,1)


    # ### 옵션차액 계산
    # * 기본판매가의 최소값, 최대값 추출
    price_max = goods_clear['기본판매가'].max()
    price_min = goods_clear['기본판매가'].min()

    # ### 엑셀에 적힐 기본 판매가격 계산
    # * 옵션별 판매가격이 차이가 없을 경우는 최소 금액이 판매가격이 됨
    basePrice = np.int64(round(price_min,-2))

    # * 정해 놓은 마진 이상 남도록 최종판매가 다시 계산
    # * setting시트에서 불러온 최소마진 설정값과 1차 계산 시 도출된 마진의 최소값과 비교한다.
    # * 마진 리스트의 최소값이 < 최소마진(marginMin) 일 때 부족한 만큼 판매가격을 높여준다.

    if marginMin > goods_clear['마진'].min():
        price_correction = round((marginMin-goods_clear['마진'].min())*1.15)
        price_correction = np.int64(round(price_correction,-2))

    else :
        price_correction = 0
        price_correction = np.int64(round(price_correction,-2))
        
    # * 최종판매가 = 기본판매가격+마진보정금액
    tune_marginPrice = basePrice + price_correction

    # 표시 판매가 계산
    dp_price = np.int64(round(tune_marginPrice / (1-discount_rate/100),-2))
    goods_clear['옵션차액'] = round(goods_clear['기본판매가'] - price_min,-2)

    #할인금액 계산
    discount_price = dp_price - tune_marginPrice

    # 메모란에 경고 메시지를 찍어 줄 것임.
    warningMemoList = []
    warningMemo =""

    # * 배송비 셋팅에서 유료 배송일 경우 판매가격에서 배송비를 차감하고 배송비 필드에 배송비 셋팅값을 입력한다.
    # 옵션차액이 판매가의 50%를 넘을 경우 판매가를 재 조정한다.

    if goods_clear['옵션차액'].max() >= dp_price*0.5:
        
        if ship_method == "무료":
            errorCorrectionPrice = goods_clear['옵션차액'].max() * 2 - dp_price + 1000
            finalPrice = dp_price + errorCorrectionPrice
            print("* [배송선택] - 무료배송")
            
        elif ship_method == "유료" or "수량별":
            errorCorrectionPrice = goods_clear['옵션차액'].max() * 2 - dp_price - rship_price  + 1000
            finalPrice = dp_price + errorCorrectionPrice
            print("* [배송선택] - 유료 or 수량별 배송")

        discount_price = discount_price + errorCorrectionPrice
        finalPrice = np.int64(round(finalPrice,-2))
        warningMemoList.append(f'* 스마트스토어 규정상 옵션 차액은 판매가의 50%를 넘을 수 없습니다.\n* 업로드 오류 방지를 위해 판매가와 할인가를 조정했습니다.\n* 실제 고객의 결제금액 및 마진과는 상관없음\n* [보정금액]: 판매가와 할인가에 각각 +{errorCorrectionPrice}원]')
        print(Fore.YELLOW + '* 옵션차액이 50%가 넘는 옵션이 존재하여 판매가와 할인가를 조정하였습니다. 자세한 내용은 엑셀 시트의 <메모>란 확인 하세요. '+Fore.RESET)
        

    else:
        if ship_method == "유료" or "수량별":
            print("* [배송선택] - 유료배송")
            finalPrice = dp_price-rship_price
            finalPrice = np.int64(round(finalPrice,-2))
        else:
            print("* [배송선택] - 유료 or 수량별 배송")
            finalPrice = dp_price
            finalPrice = np.int64(round(finalPrice,-2))

    tuneMargin = round(tune_marginPrice-goods_clear['구매원가'].min()-(tune_marginPrice*fee_naver/100),-2)
    tuneMarginRate = round(tuneMargin/tune_marginPrice*100,0)

    if duty >=150:
        warningMemoList.append( '* 옵션에 관부가세 대상이 되는 $150이상 품목이 있습니다. 소싱 금액을 점검하세요.\n')
        print(Fore.YELLOW+ '메모란 확인 - $150이상 품목있음. 관부가세주의'+Fore.RESET)
    else:
        pass

    warningMemo = str("\n".join(warningMemoList))

    print('6. 판매 가격 계산 완료!')

    # ### 옵션항목 뽑기
    #optionT1, T2는 옵션항목의 필드명들 순서대로 표기
    #입력가격이 몇 번째 옵션을 기준하여 작성되었는지 판정
    #옵션1, 옵션2 칸에 입력한 옵션내용 데이터를 중복제거하고 데이터프레임을 필터링해본다.
    #필터링된 데이터프레임에서 옵션차액 필드 내의 데이터의 중복 검사를 해본다.
    #중복검사에서 1개가 나오면 모든 데이터가 1개의 가격으로 쓰여져 있다는 뜻..그러므로 주요한 가격이 아님
    #2개 이상이 나오면 그 필터링된 제목을 가진 옵션이 가격을 결정하는 것
    #그 제목의 순서대로 옵션차액을 기록한다.
    option_list1 = []
    option_list2 = []
    option_list3 = []
    opPrice_list = []

    df_gc = goods_clear.astype(str)
    deff_price = ""
    optionPrice = ""
    deff_list = []
    zerodeff_list = []

    if optionColcnt == 6:
        df_gcprice = df_gc.drop_duplicates(subset=optionT1,ignore_index=False)
        df_subset1 = df_gcprice['옵션차액'].drop_duplicates()
        dupPriceCnt1 = df_subset1.value_counts().sum(axis=0) #T1열의 구성 데이터의 중복 개수
        quanty_list = list(df_gcprice['재고수량'])
        quantyString = ",".join(quanty_list) #재고수량을 naver 포멧으로 변경

        df_gcprice2 = df_gc.drop_duplicates(subset=optionT2,ignore_index=False)
        df_subset2 = df_gcprice2['옵션차액'].drop_duplicates()
        dupPriceCnt2 = df_subset2.value_counts().sum(axis=0) #T2열의 구성 데이터의 중복 개수

        if dupPriceCnt1 >= 2:
            print("* [입력옵션] - 첫번째 옵션을 '주 옵션'으로 가격을 계산 합니다.")
            df_option1 = df_gc[optionT1].drop_duplicates()
                    
            for op in df_option1:
                option_deff = goods_clear.loc[goods_clear[optionT1] == op]
                intdeff = option_deff['옵션차액'].drop_duplicates()

                strdeff = np.int64(intdeff.min())
                deff_list.append(strdeff)

            deff_list = list(map(str, deff_list))
            zerodeff_list.extend(["0"] * len(deff_list))



            deff_price = str(",".join(deff_list))
            zero_deff = str(",".join(zerodeff_list))
            optionPrice = deff_price + '\n' + zero_deff # optionPrice

        elif dupPriceCnt2 >= 2:
            print("* [입력옵션] - 두번째 옵션을 '주 옵션'으로 가격을 계산 합니다.")
            df_option1 = df_gc[optionT2].drop_duplicates()
            
            for op in df_option1:
                option_deff = goods_clear.loc[goods_clear[optionT2] == op]
                intdeff = option_deff['옵션차액'].drop_duplicates()
                strdeff = np.int64(intdeff.min())
                deff_list.append(strdeff)

            deff_list = list(map(str, deff_list))
            zerodeff_list.extend(["0"] * len(deff_list))

            deff_price = str(",".join(deff_list))
            zero_deff = str(",".join(zerodeff_list))
            optionPrice = zero_deff+'\n'+ deff_price

        else:
            print('* [입력옵션] - 옵션의 가격이 모두 동일합니다.')
            df_option1 = df_gc[optionT1].drop_duplicates()
            for op in df_option1:
                option_deff = goods_clear.loc[goods_clear[optionT1] == op]
                intdeff = option_deff['옵션차액'].drop_duplicates()
                strdeff = np.int64(intdeff.min())
                deff_list.append(strdeff)
            
            deff_list = list(map(str, deff_list))
            zerodeff_list.extend(["0"] * len(deff_list))
            deff_price = str(",".join(deff_list))
            zero_deff = str(",".join(zerodeff_list))
            optionPrice = deff_price
            
    elif optionColcnt == 5:
        df_gcprice = df_gc.drop_duplicates(subset=optionT1,ignore_index=False)
        df_option1 = df_gc[optionT1].drop_duplicates()  # 첫번째 필드의 데이터들을 프레임에 담는다.
        quanty_list = list(df_gcprice['재고수량'])
        quantyString = ",".join(quanty_list) #재고수량을 naver 포멧으로 변경

        # 일단 같은 옵션명과 금액을 가진 놈들을 뽑아 중복제거 후 리스트에 담는다.
        # int로 변경 후 다른 이름을 또 검색해서 중복제거 후 계속 추가한다.
        # 완성된 리스트를 스트링으로 변환한다.
        for op in df_option1:
            option_deff = goods_clear.loc[goods_clear[optionT1] == op]
            intdeff = option_deff['옵션차액'].drop_duplicates()
            strdeff = np.int64(intdeff.min())
            deff_list.append(strdeff)
            
        # join 함수를 사용할 때는 리스트 내의 인자들이 모두 string 형태여야 한다. 그러니깐.
        deff_list = list(map(str, deff_list))
        deff_price = ",".join(deff_list)
        optionPrice = deff_price
        
    # 네이버가 요구하는 양식으로 데이터를 편집하여 스트링으로 저장

    if optionColcnt == 6:
        df_option1 = goods_clear[optionT1].drop_duplicates()
        df_option2 = goods_clear[optionT2].drop_duplicates()
        list_option1 = df_option1.values.tolist()  # 담겨진 데이터들 중 중복 삭제하고 유일한 값들만 모아서 프레임에 저장
        list_option2 = df_option2.values.tolist()
        optionDesc1 = ",".join(map(str,list_option1))
        optionDesc2 = ",".join(map(str,list_option2))
        optionValue = optionDesc1 + '\n' + optionDesc2

    elif optionColcnt == 5:
        df_option1 = goods_clear[optionT1].drop_duplicates()
        list_option1 = df_option1.values.tolist()
        optionDesc1 = ",".join(map(str,list_option1))
        optionValue = optionDesc1

    print('7. 옵션 작성 완료!')
    return finalPrice, discount_price, optionValue, optionPrice, warningMemo, rate, currency_type, prime_cost, tune_marginPrice, tuneMargin, tuneMarginRate, quantyString

def make_html(desc_html, pName, addDescBool, opImg_position): #상세페이지 작성 기능

    try:
        preDescPages = desc_html
        descPages2 = re.sub("img referrerpolicy='no-referrer'|{LINK}|", "", preDescPages)
        descPages1 = re.sub("< ", "<", descPages2)+'\n'
        descPages = '<div align="center"><!-- 상세페이지 수정은 여기서부터 -->' + descPages1 + '<!-- 상세페이지 수정은 여기까지 --></div>'

    except TypeError:
        print(Fore.RED + '오류 - product.xlsx->상세페이지 필드에 url이 없거나 잘못 되었습니다.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
    

    descPname = '<br><br><h1 style="text-align: center;"><strong>' + pName + "</strong></h1><br><br>"+'<br>'
    naverTop = '<br>'+'<div align="center"><!-- 여기서부터 상단 공지 이미지 --><img src="' + naver_top + '"/></div>'+'<br>'
    naverBottom ='<br>'+ '<div align="center"><!-- 여기서부터 하단 공지1 이미지 --><img src="' + naver_bottom + '"/></div>'+'<br>'
    naverBottom2 ='<br>'+ '<div align="center"><!-- 여기서부터 하단 공지2 이미지 --><img src="' + naver_bottom2 + '"/></div>'
    #shop11Top = '<img src="' + shop11st_top + '"/>'+'\n'
    #shop11stBottom = '<img src="' + shop11st_bottom + '"/>'+'\n'
    
    df_optiongoods = writeSheet_DF.iloc[0:,5:11]
    df_optiongoods.replace('', np.nan, inplace=True)
    goods_clear = df_optiongoods.dropna(axis=1).copy()
    option_gooddf = goods_clear.columns
    optionColcnt = len(goods_clear.columns)

    optionT1 = []
    optionT2 = []

    if optionColcnt == 6:
        optionT1 = option_gooddf[0]
        optionT2 = option_gooddf[1]

    elif optionColcnt == 5:
        optionT1 = option_gooddf[0]
    try:
        df_opurl = writeSheet_DF.iloc[0:,4:7]
        df_filter = df_opurl.drop_duplicates(subset=optionT1,ignore_index=False)
        img_option = df_filter['옵션이미지']
        img_optionTag = img_option.str.replace(r"<img\s+src=['\"]", '', regex=True)
        img_optionTag = img_optionTag.str.replace(r"['\"]\s*/?>", '', regex=True)
        op_imgurls = img_optionTag.str.replace(r"\n", '', regex=True).tolist()
    except KeyError:
        print(Fore.RED + '오류 - 옵션이미지 필드에 url이 없거나 잘못 되었습니다.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
    except AttributeError:
        print(Fore.RED + '오류 - 옵션이미지 url이 잘못 되었을 수 있으나 그대로 다시 실행해 보세요.')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
        
    OpTitle = df_filter[optionT1]
    op_titlelist = OpTitle.values.tolist()
    optionLen = len(op_titlelist)

    opjoin_list = []
    cntj=1

    for i in range(optionLen):
        
        try :
            strtitle = '<div><h2><strong>옵션'+str(cntj)+'. '+ op_titlelist[i]+'</strong></h2></div>'
            strImg = '<div align="center"><img src="'+op_imgurls[i]+'"/></div><br><br>'
            opjoin_list.append(strtitle+strImg)
            cntj += 1
        
        except TypeError as e:
            print(Fore.RED + '오류 - 옵션 url을 입력하지 않은 것 같습니다.')
            print(Fore.RESET + "엔터를 누르면 종료합니다.")
            aInput = input("")
            sys.exit()
        
    opjoinStr = str("\n".join(opjoin_list))
    optionHtml = '<br><!-- 옵션 이미지 시작 --><div align="center"><div align="center"><img src="https://ai.esmplus.com/letsbuying/notice/option-Img.png" alt="option-Img" border="0"></div><br>' + opjoinStr + '<!-- 옵션 이미지 끝 -->'

    descNaver = ""
    p_desc = ""
    #desc11st = ""

    if addDescBool == 0:
        if opImg_position == 0:
            descNaver = naverTop + descPname + descPages + optionHtml + naverBottom + naverBottom2
            p_desc = descPname + descPages + optionHtml
        elif opImg_position == 1:
            descNaver = naverTop + descPname + optionHtml + descPages + naverBottom + naverBottom2
            p_desc = descPname + optionHtml + descPages
        descNaver = descNaver.replace('<img src=""/>', '')
        descPN = "<div align='center'>" + descNaver + "</div>"
        descSharing = "<div align='center'>" + p_desc + "</div>"

    elif addDescBool == 1:
        if opImg_position == 0:
            descNaver = descPname + descPages + optionHtml
            p_desc = descPname + descPages + optionHtml
        elif opImg_position == 1:
            descNaver = descPname + optionHtml + descPages
            p_desc = descPname + optionHtml + descPages
            
        descNaver = descNaver.replace('<img src=""/>', '')
        descPN = "<div align='center'>" + p_desc + "</div>"
        descSharing = "<div align='center'>" + p_desc + "</div>"
        
    else:
        print(Fore.RED + "오류 - 상하단 이미지 등록 여부가 잘못 입력 되었습니다." + Fore.RESET+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")

    print("8. 상세페이지 작성 완료!")
    return descPN, descSharing, op_imgurls, descPages

def veiw_Desc(descPN): # 미리 보기 버튼 클릭 시 html 파일을 생성하여 브라우저에서 실행 시킴
    current_dir = os.getcwd()
    
    file_name = '/temp.html'
    file_path = current_dir + file_name
    print(file_path)
    with open(file_path, 'w') as file:
        file.write(descPN)
    webbrowser.open_new_tab(file_path)
    print('브라우저 실행!')


password = loadPassword() #set.ini 파일에서 패스워드를 읽는 함수
passTag = getPtag("https://sites.google.com/view/test-exceldoc/pass") #관리자 패스워드가 저장된 웹페이지 url을 전달하여 패스워드를 크롤링 해 오는 getPtag 함수 실행
judge(password,passTag) #웹에서 가져온 패스워드와 set.ini 파일에 저장된 패스워드를 비교하여 틀리면 입력창으로 입력받고 맞으면 통과시킴

mainImg_file_path = './mainImages'
mainImg_output_path = './mainImages'
product_path = './product.xlsx' # product.xlsx 파일이 있는 경로
setting_path = './product.xlsx' # 셋팅 관련 엑셀 파일이 있는 경로. 유저가 선택할 수 없다.

########## 엑셀 파일을 오픈하면 아래 코드 바로 실행 ############
writeSheet_DF, setting_DF = readExcel(product_path, setting_path) #엑셀 파일을 불러서 데이터프레임 형식으로 만든다.

# 파일저장용 시간 불러오기
tday = time.time()
fday = time.strftime('%Y%m%d',time.localtime(time.time()))
tday_s = time.strftime('%Y%m%d-%H%M%S',time.localtime(time.time()))
tday_f = time.strftime('%Y-%m-%d',time.localtime(time.time()))

# 읽어온 셋팅 데이터 프레임에서 각 셋팅변수 선언
# 계산이 필요한 금액은 숫자형으로 변경
set_list = list(setting_DF['입력값'])
nickName = set_list[0]  #닉네임
as_info = set_list[1]   #as안내 내용
as_tel = set_list[2]    #A/s전화번호
factory_desc = set_list[3]  #제조사
brand_info = set_list[4]    #브랜드
discount_rate = float(set_list[5]) #표시 될 할인율
ship_method = set_list[6]   #배송비유형
qt_charge = set_list[7]     #수량별부과-수량
rship_price = int(set_list[8])   #기본배송비
check_method = set_list[9]  #배송비 결제방식
refund_ship = set_list[10]   #반품배송비
exchange_ship = set_list[11] #교환배송비
gift_desc = set_list[12] #사은품
point_tReview = set_list[13]    #텍스트리뷰 작성시 지급 포인트
point_photoReview = set_list[14]    #포토/동영상 리뷰 작성시 지급 포인트
point_monthText = set_list[15]  #한달사용 텍스트리뷰 작성시 지급 포인트
point_monthVideo = set_list[16] #한달사용 포토/동영상리뷰 작성시 지급 포인트
point_talktalk = set_list[17]   #톡톡친구/스토어찜고객 리뷰 작성시 지급 포인트
rate_CNY = float(set_list[18]) #환율CNY
rate_USD = float(set_list[19]) #환율USD
fomul = float(set_list[20])    #가격조정값
fee_naver = float(set_list[21])    #네이버수수료
marginMin = int(set_list[22])    #최소마진
naver_top = str(set_list[23])    #스스 상세페이지에 삽입되는 상단이미지
naver_bottom = str(set_list[24]) #스스 상세페이지에 삽입되는 하단이미지 1
naver_bottom2 = str(set_list[25]) #스스 상세페이지에 삽입되는 하단이미지 2
addDescBool = int(set_list[26])  #개인 상세페이지 상,하단 이미지 사용 유무
opImg_position = int(set_list[27]) #옵션이미지 위치

# * product.xlsx 파일->wirte 시트에서 유저가 입력한 값 추출
shop_type =writeSheet_DF['사이트'][0] ### wirte 시트 url 필드에서 쇼핑몰 종류별로 상품ID 추출 및 표준url 생성
url_shop = writeSheet_DF['url'][0]
productCord, product_url = extract_id(shop_type, url_shop) #productCord, product_url을 추출하는 함수 실행
pName = writeSheet_DF['상품명'][0]
categori_num = int(writeSheet_DF['카테고리번호'][0])
desc_html = writeSheet_DF['상세페이지'][0] #상세페이지 추출
videourl = str(writeSheet_DF['동영상url'][0]) # 비디오url 추출
optionTitle = optionTitle(writeSheet_DF) #옵션명 추출하여 네이버 포멧으로 변경
progress_text(productCord, videourl) #위의 추출된 데이터의 결과 텍스트 출력

#추출한 카테고리 번호로 네이버 카테고리 전체 경로 이름을 찾아서 기입한다.
strCalevel1, strCalevel2, strCalevel3, strCalevel4 = write_categori_name(categori_num)

# gui상에 표시될 카테고리 텍스트
fullname_categori = strCalevel1 +' > '+ strCalevel2 +' > '+ strCalevel3 +' > '+ strCalevel4

#모든 가격 계산
finalPrice, discount_price, optionValue, optionPrice, warningMemo, rate, currency_type, prime_cost, tune_marginPrice, tuneMargin, tuneMarginRate, quantyString = price_Calculation(writeSheet_DF)

########## 여기까지 ############

# 1. 이미지 파일명을 불러옴. 2. 이미지 파일명을 변경함. 3. 이미지 파일 저장소로 복사
mainImage, subImages = mainImg_Edit(mainImg_file_path, mainImg_output_path)

#상세페이지 미리보기 버튼 클릭시
#상세페이지 작성 후 브라우저로 보여줌
descPN, descSharing, op_imgurls, descPages = make_html(desc_html, pName, addDescBool, opImg_position)
veiw_Desc(descPN)


# ### 엑셀에 기재될 배송비
if ship_method == "유료" or "수량별":
    ship_price = rship_price
else:
    ship_price = 0

ship_price = str(ship_price)

#스마트스토어 필드명 불러오기
store_field = pd.read_excel('./product.xlsx', sheet_name = 'store', header = 0)
storeField_list = list(store_field['네이버'])

#스마트스토어 본인용 엑셀파일 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.append(storeField_list)

#스마트스토어 배포용 엑셀파일 생성
p_wb = openpyxl.Workbook()
p_ws = p_wb.active
p_ws.append(storeField_list)

#스마트스토어 본인용 엑셀파일 작성
ws["A2"].value = "신상품"
ws["B2"].value = categori_num
ws["C2"].value = pName
ws["D2"].value = finalPrice
ws["E2"].value = "999"
ws["F2"].value = as_info
ws["G2"].value = as_tel
ws["H2"].value = mainImage
ws["I2"].value = subImages
ws["J2"].value = descPN
ws["k2"].value = productCord
ws["L2"].value = " "
ws["M2"].value = factory_desc
ws["N2"].value = brand_info
ws["O2"].value = " "
ws["P2"].value = " "
ws["Q2"].value = "과세상품"
ws["R2"].value = "Y"
ws["S2"].value = "Y"
ws["T2"].number_format = '"0"#'
cellFormat = ws["T2"]
cellFormat.number_format = '@'
ws["T2"].value = "0200037"
ws["U2"].value = factory_desc
ws["V2"].value = "N"
ws["W2"].value = " "
ws["X2"].value = "택배, 소포, 등기"
ws["Y2"].value = "CJGLS"
ws["Z2"].value = ship_method
ws["AA2"].value = rship_price
ws["AB2"].value = check_method
ws["AC2"].value = " "
ws["AD2"].value = qt_charge
ws["AE2"].value = refund_ship
ws["AF2"].value = exchange_ship
ws["AG2"].value = " "
ws["AH2"].value = " "
ws["AI2"].value = " "
ws["AJ2"].value = discount_price
ws["AK2"].value = "원"
ws["AL2"].value = " "
ws["AM2"].value = " "
ws["AN2"].value = " "
ws["AO2"].value = " "
ws["AP2"].value = " "
ws["AQ2"].value = " "
ws["AR2"].value = point_tReview
ws["AS2"].value = point_photoReview
ws["AT2"].value = point_monthText
ws["AU2"].value = point_monthVideo
ws["AV2"].value = point_talktalk
ws["AW2"].value = " "
ws["AX2"].value = gift_desc
ws["AY2"].value = "조합형"
ws["AZ2"].value = optionTitle
ws["BA2"].value = optionValue
ws["BB2"].value = optionPrice
ws["BC2"].value = quantyString
ws["BD2"].value = " "
ws["BE2"].value = " "
ws["BF2"].value = " "
ws["BG2"].value = " "
ws["BH2"].value = "상세페이지 참조"
ws["BI2"].value = "상세페이지 참조"
ws["BJ2"].value = "상세페이지 참조"
ws["BK2"].value = "상세페이지 참조"
ws["BL2"].value = "N"
ws["BM2"].value = " "
ws["BN2"].value = " "
ws["BO2"].value = " "
ws["BP2"].value = " "
ws["BQ2"].value = " "
ws["BR2"].value = " "
ws["BS2"].value = " "
ws["BT2"].value = " "
ws["BU2"].value = " "
ws["BV2"].value = warningMemo
ws["BV2"].font = Font(color="FF0000")
ws["BW2"].value = nickName # 작성자
ws["BX2"].value = tday_f # 소싱일
ws["By2"].value = shop_type #소싱사이트
ws["Bz2"].value = productCord #판매자상품코드
ws["CA2"].value = pName #제품명
ws["CB2"].value = product_url #제품URL
ws["CC2"].value = writeSheet_DF['물건가격'].min()
ws["CD2"].value = rate #적용환율
ws["CE2"].value = currency_type #결제통화
ws["CF2"].value = writeSheet_DF['실제배송비'].min()
ws["CG2"].value = round(prime_cost,-2)
ws["CH2"].value = round(tune_marginPrice,-2)
ws["CI2"].value = round(tuneMargin,1)
ws["CJ2"].value = round(tuneMarginRate,1)
ws["CK2"].value = fomul
ws["CL2"].value = marginMin
ws["CM2"].value = categori_num
ws["CN2"].value = strCalevel1
ws["CO2"].value = strCalevel2
ws["CP2"].value = strCalevel3
ws["CQ2"].value = strCalevel4

#스마트스토어 본인용 엑셀파일 작성
p_ws["A2"].value = "신상품"
p_ws["B2"].value = categori_num
p_ws["C2"].value = pName
p_ws["D2"].value = finalPrice
p_ws["E2"].value = "999"
p_ws["F2"].value = "as_info"
p_ws["G2"].value = "000-000-0000"
p_ws["H2"].value = mainImage
p_ws["I2"].value = subImages
p_ws["J2"].value = descSharing
p_ws["k2"].value = productCord
p_ws["L2"].value = " "
p_ws["M2"].value = "factory_desc"
p_ws["N2"].value = "brand_info"
p_ws["O2"].value = " "
p_ws["P2"].value = " "
p_ws["Q2"].value = "과세상품"
p_ws["R2"].value = "Y"
p_ws["S2"].value = "Y"
cellFormat = p_ws["T2"]
cellFormat.number_format = '@'
p_ws["T2"].value = "0200037"
p_ws["U2"].value = "factory_desc"
p_ws["V2"].value = "N"
p_ws["W2"].value = " "
p_ws["X2"].value = "택배, 소포, 등기"
p_ws["Y2"].value = "CJGLS"
p_ws["Z2"].value = ship_method
p_ws["AA2"].value = rship_price
p_ws["AB2"].value = check_method
p_ws["AC2"].value = " "
p_ws["AD2"].value = qt_charge
p_ws["AE2"].value = refund_ship
p_ws["AF2"].value = exchange_ship
p_ws["AG2"].value = " "
p_ws["AH2"].value = " "
p_ws["AI2"].value = " "
p_ws["AJ2"].value = discount_price
p_ws["AK2"].value = "원"
p_ws["AL2"].value = " "
p_ws["AM2"].value = " "
p_ws["AN2"].value = " "
p_ws["AO2"].value = " "
p_ws["AP2"].value = " "
p_ws["AQ2"].value = " "
p_ws["AR2"].value = point_tReview
p_ws["AS2"].value = point_photoReview
p_ws["AT2"].value = point_monthText
p_ws["AU2"].value = point_monthVideo
p_ws["AV2"].value = point_talktalk
p_ws["AW2"].value = " "
p_ws["AX2"].value = gift_desc
p_ws["AY2"].value = "조합형"
p_ws["AZ2"].value = optionTitle
p_ws["BA2"].value = optionValue
p_ws["BB2"].value = optionPrice
p_ws["BC2"].value = quantyString
p_ws["BD2"].value = " "
p_ws["BE2"].value = " "
p_ws["BF2"].value = " "
p_ws["BG2"].value = " "
p_ws["BH2"].value = "상세페이지 참조"
p_ws["BI2"].value = "상세페이지 참조"
p_ws["BJ2"].value = "상세페이지 참조"
p_ws["BK2"].value = "상세페이지 참조"
p_ws["BL2"].value = "N"
p_ws["BM2"].value = " "
p_ws["BN2"].value = " "
p_ws["BO2"].value = " "
p_ws["BP2"].value = " "
p_ws["BQ2"].value = " "
p_ws["BR2"].value = " "
p_ws["BS2"].value = " "
p_ws["BT2"].value = " "
p_ws["BU2"].value = " "
p_ws["BV2"].value = warningMemo
p_ws["BV2"].font = Font(color="FF0000")
p_ws["BW2"].value = nickName # 작성자
p_ws["BX2"].value = tday_f # 소싱일
p_ws["By2"].value = shop_type #소싱사이트
p_ws["Bz2"].value = productCord #판매자상품코드
p_ws["CA2"].value = pName #제품명
p_ws["CB2"].value = product_url #제품URL
p_ws["CC2"].value = writeSheet_DF['물건가격'].min()
p_ws["CD2"].value = rate #적용환율
p_ws["CE2"].value = currency_type #결제통화
p_ws["CF2"].value = writeSheet_DF['실제배송비'].min()
p_ws["CG2"].value = round(prime_cost,-2)
p_ws["CH2"].value = round(tune_marginPrice,-2)
p_ws["CI2"].value = round(tuneMargin,1)
p_ws["CJ2"].value = round(tuneMarginRate,1)
p_ws["CK2"].value = fomul
p_ws["CL2"].value = marginMin
p_ws["CM2"].value = categori_num
p_ws["CN2"].value = strCalevel1
p_ws["CO2"].value = strCalevel2
p_ws["CP2"].value = strCalevel3
p_ws["CQ2"].value = strCalevel4

#엑셀 파일 저장
new_fileName = ('./excel/'+productCord+'_'+'개인용'+'_'+tday_s+'.xlsx')
wb.save(new_fileName)
print("10. 개인용파일 작성완료!")

new_fileName = ('./excel/'+productCord+'_'+'배포용'+'_'+tday_s+'.xlsx')
p_wb.save(new_fileName)
print("11. 배포용파일 작성완료!")

# 이미지 저장용 경로 설정
path_productCord = ""
path_productCord = './excel/'+ productCord
path_Desc = './excel/'+ productCord +'/Desc'
path_Option = './excel/'+ productCord +'/Option'
path_Backup = './excel/product_backup'

# 다운로드 이미지 저장용 폴더 생성
createFolder(path_productCord)
createFolder(path_Desc)
createFolder(path_Option)
createFolder(path_Backup)
print('12. 다운로드 폴더 생성 완료!'+'\n')

opImg_Download(op_imgurls) # 옵션 이미지 다운로드
descImg_Download(descPages) # 상세페이지 이미지 다운로드

fVideoUrl = open('./excel/' + productCord + '/동영상주소.txt','w') #video url을 저장할 텍스트 파일 생성
fVideoUrl.write(videourl) #video url을 텍스트파일에 쓰기
fVideoUrl.close() #텍스트 파일 닫기

copy_df = writeSheet_DF #백업 파일 생성
copy_df = writeSheet_DF.to_excel(excel_writer=path_Backup+'/product_'+productCord+'_'+tday_s+'.xlsx', index=False) #백업 파일 저장

print('\n'+ Fore.LIGHTBLUE_EX + "완성! 엔터를 누르면 종료합니다." + Fore.RESET)
aInput = input("")